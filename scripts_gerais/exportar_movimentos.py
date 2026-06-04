# exportar_movimentos.py
# Extrai todos os documentos da coleção `movimentos` do Firestore
# e gera um arquivo Excel com todas as informações.
#
# Uso:
#   python exportar_movimentos.py
#
# Saída: movimentos_export.xlsx (mesmo diretório do script)

import os
import json
from datetime import datetime

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Firebase
# ─────────────────────────────────────────────────────────────────────────────

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_CRED_PATH = os.path.join(_THIS_DIR, "motiva-8b82f-firebase-adminsdk-fbsvc-14d8d2b5e8.json")
_PROJECT_ID = "motiva-8b82f"

if not firebase_admin._apps:
    try:
        firebase_admin.initialize_app(
            credentials.ApplicationDefault(),
            {"projectId": _PROJECT_ID},
        )
    except Exception:
        firebase_admin.initialize_app(credentials.Certificate(_CRED_PATH))

db = firestore.client()

# ─────────────────────────────────────────────────────────────────────────────
# Extração
# ─────────────────────────────────────────────────────────────────────────────

def fetch_movimentos():
    docs = db.collection("movimentos").stream()
    rows = []
    for d in docs:
        data = d.to_dict() or {}
        rows.append({
            "id":               d.id,
            "displayName":      data.get("displayName") or data.get("name", ""),
            "categories":       ", ".join(data.get("categories") or []),
            "primaryMuscles":   ", ".join(data.get("primaryMuscles") or []),
            "equipment":        ", ".join(data.get("equipment") or []),
            "prType":           data.get("prType", ""),
            "supportedPrTypes": ", ".join(data.get("supportedPrTypes") or []),
            "prUnit":           data.get("prUnit", ""),
        })
    rows.sort(key=lambda r: (r["categories"], r["displayName"].lower()))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    ("ID",                 "id",               20),
    ("Nome",               "displayName",      28),
    ("Categorias",         "categories",       22),
    ("Músculos primários", "primaryMuscles",   30),
    ("Equipamento",        "equipment",        22),
    ("Tipo de PR",         "prType",           14),
    ("Tipos suportados",   "supportedPrTypes", 22),
    ("Unidade PR",         "prUnit",           13),
]

# Cores por categoria (fundo da linha)
_CAT_COLORS = {
    "levantamento de peso": "EBF5FB",
    "ginástica":            "EAFAF1",
    "metcon":               "FEF9E7",
    "força":                "F9EBF8",
    "habilidade":           "FDF2F8",
}
_DEFAULT_COLOR = "F8F9FA"

_HEADER_FILL  = PatternFill("solid", fgColor="1A3A8A")   # azul escuro
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DATA_FONT    = Font(name="Calibri", size=10)
_THIN_BORDER  = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def build_excel(rows: list, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentos"

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Dados ──────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        cat    = row["categories"].split(",")[0].strip().lower()
        hex_bg = _CAT_COLORS.get(cat, _DEFAULT_COLOR)
        fill   = PatternFill("solid", fgColor=hex_bg)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font      = _DATA_FONT
            cell.fill      = fill
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        ws.row_dimensions[row_idx].height = 18

    # ── Tabela resumo por categoria (mesma planilha, separada) ─────────────
    _add_summary(ws, rows, start_col=len(COLUMNS) + 2)

    wb.save(out_path)


def _add_summary(ws, rows: list, start_col: int):
    """Mini-tabela de contagem por categoria à direita dos dados."""
    from collections import Counter

    cat_counts = Counter()
    pr_counts  = Counter()

    for r in rows:
        for c in r["categories"].split(","):
            c = c.strip()
            if c:
                cat_counts[c] += 1
        if r["prType"]:
            pr_counts[r["prType"]] += 1

    sc = start_col
    title_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    sub_fill   = PatternFill("solid", fgColor="2E4099")

    # Tabela 1: por categoria
    ws.cell(1, sc, "Categoria").font      = title_font
    ws.cell(1, sc, "Categoria").fill      = sub_fill
    ws.cell(1, sc, "Categoria").alignment = Alignment(horizontal="center")
    ws.cell(1, sc + 1, "Qtd").font        = title_font
    ws.cell(1, sc + 1, "Qtd").fill        = sub_fill
    ws.cell(1, sc + 1, "Qtd").alignment   = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(sc)].width     = 22
    ws.column_dimensions[get_column_letter(sc + 1)].width = 8

    for i, (cat, count) in enumerate(sorted(cat_counts.items()), start=2):
        ws.cell(i, sc,     cat).font   = Font(name="Calibri", size=10)
        ws.cell(i, sc + 1, count).font = Font(name="Calibri", size=10)

    # Tabela 2: por tipo de PR (abaixo da primeira)
    offset = len(cat_counts) + 3
    ws.cell(offset, sc, "Tipo de PR").font      = title_font
    ws.cell(offset, sc, "Tipo de PR").fill      = sub_fill
    ws.cell(offset, sc, "Tipo de PR").alignment = Alignment(horizontal="center")
    ws.cell(offset, sc + 1, "Qtd").font         = title_font
    ws.cell(offset, sc + 1, "Qtd").fill         = sub_fill
    ws.cell(offset, sc + 1, "Qtd").alignment    = Alignment(horizontal="center")

    for i, (pr, count) in enumerate(sorted(pr_counts.items()), start=offset + 1):
        ws.cell(i, sc,     pr).font    = Font(name="Calibri", size=10)
        ws.cell(i, sc + 1, count).font = Font(name="Calibri", size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Buscando movimentos do Firestore...")
    rows = fetch_movimentos()
    print(f"  → {len(rows)} movimentos encontrados.")

    out_path = os.path.join(_THIS_DIR, "movimentos_export.xlsx")
    print("Gerando Excel...")
    build_excel(rows, out_path)

    print(f"\n✓ Arquivo salvo em:\n  {out_path}\n")
